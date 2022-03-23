## Imports de paquetes para que el programa funcione
from pathlib import Path
import os                                   
import pandas as pd 
from openpyxl import load_workbook
import smtplib
from email.message import EmailMessage
import pyautogui


## Ejecutar el programa desde la misma carpeta donde se encuentra el Excel "Lista_Contratistas.xlsx" y las carpetas de los meses 

def formatear_texto(texto): ##Funcion formatear el texto del excel (convertir a string, quitar espacios y poner todo en mayuscula )
    texto_format = str(texto).strip().upper()
    return texto_format
def get_mes(mes):  ## Funcion para formatear el mes a como estar organizado en las carpetas 
    mes = str(mes).strip().upper()
    dic_mes = { 
                "ENERO"     : "1. ENERO", 
                "FEBRERO"   : "2. FEBRERO", 
                "MARZO"     : "3. MARZO",
                "ABRIL"     : "4. ABRIL",
                "MAYO"      : "5. MAYO",
                "JUNIO"     : "6. JUNIO",
                "JULIO"     : "7. JULIO", 
                "AGOSTO"    : "8. AGOSTO", 
                "SEPTIEMBRE": "9. SEPTIEMBRE",
                "OCTUBRE"   : "10. OCTUBRE",
                "NOVIEMBRE" : "11. NOVIEMBRE",
                "DICIEMBRE" : "12. DICIEMBRE",
            }
    format_mes = dic_mes.get(mes)
    return format_mes

USUARIO_EMAIL = 'soportescontratacion@eseuniversitariadelatlantico.gov.co'  ## Email con el que se envia los correos
CONTRASENA_EMAIL = 'Contratacion*2022' ##Contraseña del email (Si se cambia la contraseña del correo cambiar aqui tambien)
RUTA_LISTA_CONTRATISTAS = r"Lista_Contratistas.xlsx" ## Archivo excel de donde se saca la imformacion de los contratistas (siempre debe estar en la misma carpeta de donde corra el script/programa)

datos = pd.read_excel(RUTA_LISTA_CONTRATISTAS) #variable para trabajar la informacion del excel
wb = load_workbook(RUTA_LISTA_CONTRATISTAS)
ws = wb.worksheets[0]
j = 2

for i in datos.index:  ##Loop para leer la informacion de las columnas del excel por el nombre 
    
    nombre = formatear_texto(datos['Nombre'][i])
    email = datos['Email'][i]
    num_contrato = formatear_texto(datos['Numero de contrato'][i]) 
    mes = get_mes(datos['Mes'][i])
    enviado = datos['Enviado'][i]
    path_contratista = os.getcwd()+ "\{}".format(mes) + "\{} {}".format(num_contrato, nombre) ##Carpeta del contratista

  
    
    if enviado == False: ##Loop para enviar correos si en la columna del excel de enviados es falso
    
        msg = EmailMessage()
        msg['Subject'] = 'DOCUMENTACION CONTRATO ESE UNIVERSITARIA DEL ATLANTICO' ## Asunto del correo electronico
        msg['From'] = USUARIO_EMAIL
        msg['To'] = email.strip().lower() ##Email al cual se va a enviar el correo

        msg.set_content('Señor(a) {},\nAdjunto se encuentra su informacion para radicar su cuenta de cobro:'.format(nombre)) ## Mensaje del correo electronico
        try: ##Try/except para que el codigo siga ejecutandose aun si el path_contratista no existe, si existe manda el correo si no sigue al siguiente contratista
            files = os.listdir(path_contratista) 
            for file in files: ## loop para adjuntar los archivos de la carpeta del contratista al correo
                path_archivo = path_contratista + "\{}".format(str(file))
                if (str(file) == "APROBACION.pdf" ) or (str(file) == "CDP.pdf" ) or (str(file) == "CONTRATO.pdf" ) or (str(file) == "RP.pdf" ) : ## IF para atachar al correo solo los documentos con estos nombres
                    with open(path_archivo, "rb") as f:
                         file_data = f.read()
                         file_name = str(file)
                    msg.add_attachment(file_data, maintype = "application", subtype = "octet-stream", filename = file_name)
            with smtplib.SMTP('smtp.office365.com', 587) as smtp: ## Envio del correo
                smtp.ehlo()
                smtp.starttls()
                smtp.login(USUARIO_EMAIL, CONTRASENA_EMAIL)
                smtp.send_message(msg)
                
            ws["E{}".format(str(j))] = "SI" ##Cambia la columna del excel Enviado de FALSO a SI
            wb.save(RUTA_LISTA_CONTRATISTAS) 
        except: 
            ws["E{}".format(str(j))] = "CARPETA NO ENCONTRADA" ##Cambia la columna del excel a carpeta no encontrada
            wb.save(RUTA_LISTA_CONTRATISTAS) 
            j = j+1
            continue
    j = j + 1
pyautogui.alert('Todos los correos han sido enviados', "Envio completado")

## PARA GUARDAR EL ARCHIVO COMO ".exe" CORRER COMANDO "pyinstaller --onefile --noconsole Correos_Contratacion.py"
## PRIMERO INSTALAR pyinstaller con "pip install pyinstaller"
        


            
        